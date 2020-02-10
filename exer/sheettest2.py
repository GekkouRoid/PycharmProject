import openpyxl as xl
from openpyxl.styles import PatternFill
from PIL import Image
import webcolors


img = Image.open('mount.jpg')
maxRow = img.height
maxCol = img.width
if maxRow > 676:
    maxRow = 675
if maxCol > 676:
    maxCol = 675

print(maxCol, maxRow)

wb = xl.load_workbook('new1.xlsx')
sheet = wb['Sheet1']
# ce = sheet.cell(1, 1)
# print(ce.value)
for i in range(1, maxRow + 1):
    for j in range(1, maxCol + 1):
        pixelColor = img.getpixel((i - 1, j - 1))
        rgb = webcolors.rgb_to_hex(pixelColor)
        strrgb = str(rgb)
        argbrgb = '00' + strrgb[1:]
        print(argbrgb)
        sheet.cell(j, i).fill = PatternFill(fgColor=argbrgb, bgColor=argbrgb, fill_type="solid")


wb.save('new1.xlsx')
wb.close()





