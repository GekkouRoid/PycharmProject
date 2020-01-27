import openpyxl as xl
from openpyxl.styles import PatternFill
from PIL import Image
import webcolors


img = Image.open('youare.png')
maxRow = img.height
maxCol = img.width

print(maxCol, maxRow)

wb = xl.load_workbook('new.xlsx')
sheet = wb['Sheet1']
# ce = sheet.cell(1, 1)
# print(ce.value)
for i in range(1, maxRow + 1):
    for j in range(1, maxCol +1):
        pixelColor = img.getpixel((i - 1, j -1))
        rgb = webcolors.rgb_to_hex(pixelColor)
        strrgb = str(rgb)
        argbrgb = '00' + strrgb[1:]

        sheet.cell(j, i).fill = PatternFill(fgColor=argbrgb, bgColor=argbrgb, fill_type="solid")


wb.save('new.xlsx')
wb.close()








