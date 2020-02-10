import openpyxl as xl
from openpyxl.styles import PatternFill
import webcolors

"""
This is to test the method to fill 
the cells that one unit could move to.

"""

draw_color = '006187b4'
wb = xl.load_workbook('move.xlsx')
sheet = wb['Sheet1']

move_ability = 8

x_init = 1
y_init = 1

unit_x = 13
unit_y = 14

x_move_range = [0]
y_move_range = [0]

cells_to_draw = []
count_of_cells = 0

for x in range(-move_ability, move_ability + 1):
    if x not in x_move_range:
        x_move_range.append(x)

for y in range(-move_ability, move_ability + 1):
    if y not in y_move_range:
        y_move_range.append(y)

print(x_move_range, y_move_range)

for x in x_move_range:
    for y in y_move_range:
        if abs(y) <= abs(move_ability - abs(x)):
            # As fill the cell of Excel sheet, neither x nor y could be less than 1
            if unit_x + x >= x_init and unit_y + y >= y_init:
                cells_to_draw.append([unit_x + x, unit_y + y])

print(cells_to_draw)
print(len(cells_to_draw))

# Fill the cells
for cell_to_draw in cells_to_draw:
    j = cell_to_draw[0]
    i = cell_to_draw[1]
    sheet.cell(j, i).fill = PatternFill(fgColor=draw_color, bgColor=draw_color, fill_type="solid")

wb.save('move.xlsx')
wb.close()
