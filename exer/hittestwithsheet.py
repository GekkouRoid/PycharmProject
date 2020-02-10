import openpyxl
import random
import statistics
from openpyxl import Workbook
from openpyxl.chart.axis import Scaling
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
    axis
)


wb = openpyxl.Workbook()
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]

to_be_hit = 70
try_hit_round = 100
try_hit_set =20
hit_rounds = []
for i in range(0, try_hit_set):
    hit_rounds.append(0)

for j in range(0, try_hit_set):
    for i in range(1, try_hit_round + 1):
        sheet.cell(row=i, column=1).value = to_be_hit
        sheet.cell(row=i, column=j + 2).value = random.randint(1, 100)
        if sheet.cell(row=i, column=j + 2).value < to_be_hit:
            hit_rounds[j] += 1

for i in range(0, try_hit_set):
    sheet.cell(row=try_hit_round + 2, column=i + 2).value = hit_rounds[i]

sheet.cell(row=try_hit_round + 3, column=2).value = statistics.mean(hit_rounds)

for i in range(0, try_hit_set):
    sheet.cell(row=try_hit_round + 4, column=i + 2).value = to_be_hit

c1 = LineChart()
v1 = Reference(sheet, min_row=try_hit_round + 2, min_col=2, max_col=try_hit_set + 1)
c1.add_data(v1, titles_from_data=False, from_rows=True)
c1.y_axis.title = "Hit on Round"

c1.y_axis.scaling = Scaling(min=0, max=100)

c2 = LineChart()
v2 = Reference(sheet, min_row=try_hit_round + 4, min_col=2, max_col=try_hit_set + 1)
c2.add_data(v2, titles_from_data=False, from_rows=True)
c2.y_axis.title = "Hit Respected"
c2.y_axis.axId = 200
c2.y_axis.scaling = Scaling(min=0, max=100)
c1.y_axis.crosses = "max"
c1 += c2

sheet.add_chart(c1, 'B106')


print(hit_rounds)
wb.save('hitest.xlsx')
wb.close()

